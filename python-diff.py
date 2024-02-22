#!/usr/bin/python3

import os
import sys
import openpyxl

print(" * Welcome to the Diff Excel User System *")

vCampos=["login","Nombre","Telefono","Correo"]

ruta_ayer = sys.argv[1]
ruta_hoy = sys.argv[2]

print(" - Viejo fue : "+str(ruta_ayer))
print(" - Nuevo fue : "+str(ruta_hoy))

wb_ayer = openpyxl.load_workbook(ruta_ayer)
ws_ayer = wb_ayer.active

wb_hoy = openpyxl.load_workbook(ruta_hoy)
ws_hoy = wb_hoy.active

if os.path.exists("meta-script.sh"):
    print(" * Meta Script Cleaning System ")
    os.remove("meta-script.sh")

if os.path.exists("logs.md"):
    print(" * Meta logs Cleaning System ")
    os.remove("logs.md")
    
meta_script = open("meta-script.sh",'x')
meta_script.write("#!/bin/bash\n\n")
meta_script.write("# DO NOT EDIT THIS SCRIPT \n")
meta_script.write("# IT WILL BE AUTOMAGICALLY GENERATED\n\n")
meta_script.write("# Orig "+ruta_ayer+"\n")
meta_script.write("# Dest "+ruta_hoy+"\n")

meta_logs = open("logs.md",'x')
meta_logs.write("# DO NOT EDIT THIS SCRIPT \n")
meta_logs.write("# IT WILL BE AUTOMATICALLY GENERATED\n\n")

meta_logs.write("# Cambios realizados:\n")

def nuevosContratos():
    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1
    
    while (aux_id_hoy != None):
        
        aux_id_ayer=ws_ayer.cell(row=1,column=1).value
        fila_ayer_procesada = 1
        contrato = True
        
        while (aux_id_ayer != None):
            if aux_id_ayer == ws_hoy.cell(row=fila_hoy_procesada,column=1).value : 
                contrato = False
            # Siguiente linea mecanismo 
            fila_ayer_procesada = fila_ayer_procesada + 1
            aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value
        
        if contrato :
            auxUser = ws_hoy.cell(row=fila_hoy_procesada,column=2).value
            auxFull = ws_hoy.cell(row=fila_hoy_procesada,column=3).value
            auxTel = ws_hoy.cell(row=fila_hoy_procesada,column=4).value
            auxMail = ws_hoy.cell(row=fila_hoy_procesada,column=5).value
            auxUID = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
            print(" * Contrato a "+ws_hoy.cell(row=fila_hoy_procesada,column=3).value )
            meta_script.write("useradd -m -d \"/home/"+auxUser+"\" -s \"/bin/bash\" -u "+str(auxUID)+" -c \""+auxFull+", ,"+str(auxTel)+", ,"+auxMail+"\" \""+auxUser+"\"\n" )
            meta_script.write("echo \""+auxUser+":"+str(auxTel)+"\"| chpasswd \n")
            
            meta_logs.write("\n# - Se ha creado un nuevo usuario '"+auxUser+"' con la información:\n")
            meta_logs.write(f"#   - Nombre: {auxFull}\n")
            meta_logs.write(f"#   - Teléfono: {auxTel}\n")
            meta_logs.write(f"#   - Correo: {auxMail}\n")
            meta_logs.write("useradd -m -d \"/home/"+auxUser+"\" -s \"/bin/bash\" -u "+str(auxUID)+" -c \""+auxFull+", ,"+str(auxTel)+", ,"+auxMail+"\" \""+auxUser+"\"\n")
            meta_logs.write(f"echo \"{auxUser}:{auxTel}\"| chpasswd \n")
        
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value

def modificaciones():
    meta_logs.write("\n---\n")
    meta_logs.write("# - Se han realizado modificaciones en los usuarios:\n")

    for fila_hoy_procesada in range(1, ws_hoy.max_row + 1):
        id_hoy = ws_hoy.cell(row=fila_hoy_procesada, column=1).value
        id_ayer = None

        for fila_ayer_procesada in range(1, ws_ayer.max_row + 1):
            id_ayer = ws_ayer.cell(row=fila_ayer_procesada, column=1).value

            if id_ayer == id_hoy:
                cambios = []

                for campo in range(2, 6):
                    if ws_ayer.cell(row=fila_ayer_procesada, column=campo).value != ws_hoy.cell(row=fila_hoy_procesada, column=campo).value:
                        cambios.append((campo, ws_hoy.cell(row=fila_hoy_procesada, column=campo).value))

                if cambios:
                    nombre_usuario = ws_hoy.cell(row=fila_hoy_procesada, column=2).value
                    print(f"{nombre_usuario} ha cambiado:")
                    
                    meta_logs.write("\n# - Se han realizado modificaciones en el usuario '"+nombre_usuario+"':\n")

                    for campo, nuevo_valor in cambios:
                        nombre_campo = vCampos[campo - 2]
                        print(f"  --->   {nombre_campo}: {nuevo_valor}")

                        if nombre_campo == "Nombre":
                            meta_logs.write(f"chfn -f \"{nuevo_valor}\" {nombre_usuario}\n")

                        elif nombre_campo == "login":
                            meta_logs.write(f"usermod -l {nombre_usuario} {ws_ayer.cell(row=fila_ayer_procesada, column=2).value}\n")
                        
                        elif nombre_campo == "Telefono":
                            meta_logs.write(f"chfn -r \"{nuevo_valor}\" {nombre_usuario}\n")
                        
                        elif nombre_campo == "Correo":
                            meta_logs.write(f"usermod -c \"{nuevo_valor}\" {nombre_usuario}\n")
                    
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1

def despidos():
    meta_logs.write("\n---\n")
    meta_logs.write("# - Se han realizado despidos de empleados:\n")

    for fila_ayer_procesada in range(1, ws_ayer.max_row + 1):
        id_ayer = ws_ayer.cell(row=fila_ayer_procesada, column=1).value
        id_hoy = None

        for fila_hoy_procesada in range(1, ws_hoy.max_row + 1):
            id_hoy = ws_hoy.cell(row=fila_hoy_procesada, column=1).value

            if id_ayer == id_hoy:
                break

        if id_hoy is None:
            nombre_usuario = ws_ayer.cell(row=fila_ayer_procesada, column=2).value
            print(f"{nombre_usuario} ha sido despedido.")

            meta_logs.write("\n# - Se ha eliminado el usuario '"+nombre_usuario+"':\n")
            meta_logs.write(f"deluser {nombre_usuario}\n")

nuevosContratos()
modificaciones()
despidos()

meta_script.close()
meta_logs.close()

