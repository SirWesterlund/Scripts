#!/bin/bash

# DO NOT EDIT THIS SCRIPT 
# IT WILL BE AUTOMAGICALLY GENERATED

# Orig users-240122.xlsx
# Dest users-240123.xlsx


# DESPIDOS PROCEDENTES :  

# Deleting Jesus Papi
deluser jmarchante


# NUEVOS CONTRATOS 
useradd -m -d "/home/icesar" -s "/bin/bash" -u 5006 -c "berlanas, ,452658526, ,hola@.com.ar" "icesar"
echo "icesar:452658526"| chpasswd 

# NUEVOS CONTRATOS 

sudo usermod -l apepi apepino

sudo usermod -c " apen@yo.com " apepi

sudo chfn -w " 740447599 " dtanke

sudo chfn -f Viktor Thinkk vthin
exit 0
lautaro@dns:~/Escritorio/aso$ cat python-diff.py 
#!/usr/bin/python3
### SCRIPT ACTUALIZACIÓN DE USUARIOS DE LAUTARO
import os
import sys
import openpyxl

print("\n* Welcome to the Diff Excel User System * \n")

vCampos=["login","Nombre","Telefono","Correo"]

ruta_ayer = sys.argv[1]
ruta_hoy = sys.argv[2]

print(" - Viejo fue : "+str(ruta_ayer))
print(" - Nuevo fue : "+str(ruta_hoy))

wb_ayer = openpyxl.load_workbook(ruta_ayer)
ws_ayer = wb_ayer.active

wb_hoy = openpyxl.load_workbook(ruta_hoy)
ws_hoy = wb_hoy.active

if os.path.exists("meta-script.sh"):
    print(" * Meta Script Cleaning System ")
    os.remove("meta-script.sh")
    
meta_script = open("meta-script.sh",'x')
meta_script.write("#!/bin/bash\n\n")
meta_script.write("# DO NOT EDIT THIS SCRIPT \n")
meta_script.write("# IT WILL BE AUTOMAGICALLY GENERATED\n\n")
meta_script.write("# Orig "+ruta_ayer+"\n")
meta_script.write("# Dest "+ruta_hoy+"\n")

def nuevosContratos():
    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1
    
    meta_script.write("\n# NUEVOS CONTRATOS \n")

    while (aux_id_hoy != None):
        
        aux_id_ayer=ws_ayer.cell(row=1,column=1).value
        fila_ayer_procesada = 1
        contrato = True
        
        while (aux_id_ayer != None):
            if aux_id_ayer == ws_hoy.cell(row=fila_hoy_procesada,column=1).value : 
                contrato = False
            # Siguiente linea mecanismo 
            fila_ayer_procesada = fila_ayer_procesada + 1
            aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value
        
        if contrato :
            auxUser = ws_hoy.cell(row=fila_hoy_procesada,column=2).value
            auxFull = ws_hoy.cell(row=fila_hoy_procesada,column=3).value
            auxTel = ws_hoy.cell(row=fila_hoy_procesada,column=4).value
            auxMail = ws_hoy.cell(row=fila_hoy_procesada,column=5).value
            auxUID = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
            print(" * Contrato a "+ws_hoy.cell(row=fila_hoy_procesada,column=3).value )
            
            meta_script.write("useradd -m -d \"/home/"+auxUser+"\" -s \"/bin/bash\" -u "+str(auxUID)+" -c \""+auxFull+", ,"+str(auxTel)+", ,"+auxMail+"\" \""+auxUser+"\"\n" )
            meta_script.write("echo \""+auxUser+":"+str(auxTel)+"\"| chpasswd \n")
            
        
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value

def modificaciones():

    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1
    meta_script.write("\n# NUEVOS CONTRATOS \n")

    while (aux_id_hoy != None):
        
        old_fila = 1
        old_id = ws_ayer.cell(row=old_fila,column=1).value
       
        vCambiosUsuario = []
        
        while (old_id != None):
            if old_id == aux_id_hoy:
                # fila_ayer = old_fila
                # fila_hoy es fila_hoy_procesada
                hayCambios = False
                for campo in range(2,6):
                    if ws_ayer.cell(row=old_fila,column=campo).value != ws_hoy.cell(row=fila_hoy_procesada,column=campo).value:
                        vCambiosUsuario.append(campo-2)
                        hayCambios = True
                
                if hayCambios :
                    print("\n" + ws_hoy.cell(row=fila_hoy_procesada,column=2).value + " ha cambiado : ")
                    #print(vCambiosUsuario)
                    for campoCambiado in range(0,len(vCambiosUsuario)):
                        auxC = int(vCambiosUsuario[campoCambiado])+2
                        print("  --->   " + vCampos[vCambiosUsuario[campoCambiado]] + " : " + str(ws_hoy.cell(row=fila_hoy_procesada, column=auxC).value) + "  <---")
                        campocambio = vCampos[vCambiosUsuario[campoCambiado]]
                        #print (campocambio)
                        #print ("Lo que quiero ver")
                        login = ws_hoy.cell(row=fila_hoy_procesada,column=2).value
                        if campocambio == vCampos[0]:
                            #print("cambiamos el login")
                            meta_script.write("\n")
                            meta_script.write("sudo usermod -l "  + ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+ " " +ws_ayer.cell(row=fila_hoy_procesada,column=auxC).value  +"\n")
                        elif campocambio == vCampos[1]:
                            #print("cambiamos el Nombre")
                            meta_script.write("\n")
                            meta_script.write("sudo chfn -f " + str(ws_hoy.cell(row=fila_hoy_procesada, column=auxC).value) + " " + str(login) +"\n")
                        elif campocambio == vCampos[2]:
                            #print("cambiamos el Telefono")
                            meta_script.write("\n")
                            meta_script.write("sudo chfn -w \" " + str(ws_hoy.cell(row=fila_hoy_procesada, column=auxC).value) +" \" " + login +"\n")
                        elif campocambio == vCampos[3]:
                            #print("cambiamos el Correo")
                            meta_script.write("\n")
                            meta_script.write("sudo usermod -c \" "+ str(ws_hoy.cell(row=fila_hoy_procesada, column=auxC).value)+ " \" "+login+"\n")
                
            old_fila = old_fila + 1
            old_id = ws_ayer.cell(row=old_fila,column=1).value
        
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value


def despidos():

    meta_script.write("\n\n# DESPIDOS PROCEDENTES :  \n\n")

    aux_id_ayer=ws_ayer.cell(row=1,column=1).value
    fila_ayer_procesada = 1

    while (aux_id_ayer != None):
        
        aux_id_hoy=ws_hoy.cell(row=1,column=1).value
        fila_hoy_procesada = 1
        despido = True
        
        while (aux_id_hoy != None):
            if aux_id_hoy == ws_ayer.cell(row=fila_ayer_procesada,column=1).value : 
                despido = False
            # Siguiente linea mecanismo 
            fila_hoy_procesada = fila_hoy_procesada + 1
            aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
        
        if despido :
            print(" * Despide a "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value )
            
            meta_script.write("# Deleting "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value+"\n")
            meta_script.write("deluser "+ws_ayer.cell(row=fila_ayer_procesada,column=2).value+"\n")
            meta_script.write("\n")
        
        #for columna in range(2,6):
        #    print(ws_ayer.cell(row=fila_ayer_procesada,column=columna).value)
            
        # Siguiente linea mecanismo 
        fila_ayer_procesada = fila_ayer_procesada + 1
        aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value
print ("\n\n* [DEPEDIDDOS] *\n")
despidos()
print ("\n* [CONTRATADOS] *\n")
nuevosContratos()       
print ("\n* [MODIFICADOS] *") 
modificaciones()    

# Print exit 0 
meta_script.write("exit 0\n")
meta_script.close()

sys.exit(0)
